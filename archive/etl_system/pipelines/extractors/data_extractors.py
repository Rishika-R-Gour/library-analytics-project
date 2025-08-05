#!/usr/bin/env python3
"""
Data Extraction Components
Handles various data sources for the library analytics system
"""

import pandas as pd
import sqlite3
import json
import csv
import os
from typing import Dict, List, Any, Optional
from pathlib import Path
import requests
from datetime import datetime
import openpyxl
from pipelines.etl_framework import DataExtractor, ETLPipelineError

class CSVExtractor(DataExtractor):
    """Extract data from CSV files"""
    
    def __init__(self, name: str, file_path: str, config: Dict[str, Any] = None):
        super().__init__(name, config)
        self.file_path = file_path
        self.csv_config = {
            'encoding': config.get('encoding', 'utf-8'),
            'delimiter': config.get('delimiter', ','),
            'header': config.get('header', 0),
            'skip_rows': config.get('skip_rows', None),
            'columns': config.get('columns', None)
        }
    
    def extract(self) -> pd.DataFrame:
        """Extract data from CSV file"""
        if not os.path.exists(self.file_path):
            raise ETLPipelineError(f"CSV file not found: {self.file_path}")
        
        try:
            # Read CSV with configuration
            df = pd.read_csv(
                self.file_path,
                encoding=self.csv_config['encoding'],
                delimiter=self.csv_config['delimiter'],
                header=self.csv_config['header'],
                skiprows=self.csv_config['skip_rows'],
                usecols=self.csv_config['columns']
            )
            
            self.logger.info(f"Extracted {len(df)} rows from CSV: {self.file_path}")
            return df
            
        except Exception as e:
            raise ETLPipelineError(f"Failed to read CSV file {self.file_path}: {e}")

class ExcelExtractor(DataExtractor):
    """Extract data from Excel files"""
    
    def __init__(self, name: str, file_path: str, config: Dict[str, Any] = None):
        super().__init__(name, config)
        self.file_path = file_path
        self.excel_config = {
            'sheet_name': config.get('sheet_name', 0),
            'header': config.get('header', 0),
            'skip_rows': config.get('skip_rows', None),
            'columns': config.get('columns', None),
            'engine': config.get('engine', 'openpyxl')
        }
    
    def extract(self) -> pd.DataFrame:
        """Extract data from Excel file"""
        if not os.path.exists(self.file_path):
            raise ETLPipelineError(f"Excel file not found: {self.file_path}")
        
        try:
            # Read Excel with configuration
            df = pd.read_excel(
                self.file_path,
                sheet_name=self.excel_config['sheet_name'],
                header=self.excel_config['header'],
                skiprows=self.excel_config['skip_rows'],
                usecols=self.excel_config['columns'],
                engine=self.excel_config['engine']
            )
            
            self.logger.info(f"Extracted {len(df)} rows from Excel: {self.file_path}")
            return df
            
        except Exception as e:
            raise ETLPipelineError(f"Failed to read Excel file {self.file_path}: {e}")

class DatabaseExtractor(DataExtractor):
    """Extract data from database"""
    
    def __init__(self, name: str, db_path: str, query: str, config: Dict[str, Any] = None):
        super().__init__(name, config)
        self.db_path = db_path
        self.query = query
        self.db_config = {
            'db_type': config.get('db_type', 'sqlite'),
            'connection_params': config.get('connection_params', {})
        }
    
    def extract(self) -> pd.DataFrame:
        """Extract data from database"""
        try:
            if self.db_config['db_type'] == 'sqlite':
                conn = sqlite3.connect(self.db_path)
                df = pd.read_sql_query(self.query, conn)
                conn.close()
            else:
                raise ETLPipelineError(f"Unsupported database type: {self.db_config['db_type']}")
            
            self.logger.info(f"Extracted {len(df)} rows from database query")
            return df
            
        except Exception as e:
            raise ETLPipelineError(f"Failed to extract from database: {e}")

class JSONExtractor(DataExtractor):
    """Extract data from JSON files"""
    
    def __init__(self, name: str, file_path: str, config: Dict[str, Any] = None):
        super().__init__(name, config)
        self.file_path = file_path
        self.json_config = {
            'encoding': config.get('encoding', 'utf-8'),
            'normalize_level': config.get('normalize_level', 0),
            'record_path': config.get('record_path', None)
        }
    
    def extract(self) -> pd.DataFrame:
        """Extract data from JSON file"""
        if not os.path.exists(self.file_path):
            raise ETLPipelineError(f"JSON file not found: {self.file_path}")
        
        try:
            with open(self.file_path, 'r', encoding=self.json_config['encoding']) as f:
                data = json.load(f)
            
            # Normalize JSON data to DataFrame
            if self.json_config['record_path']:
                df = pd.json_normalize(data, record_path=self.json_config['record_path'])
            else:
                df = pd.json_normalize(data)
            
            self.logger.info(f"Extracted {len(df)} rows from JSON: {self.file_path}")
            return df
            
        except Exception as e:
            raise ETLPipelineError(f"Failed to read JSON file {self.file_path}: {e}")

class APIExtractor(DataExtractor):
    """Extract data from REST API"""
    
    def __init__(self, name: str, url: str, config: Dict[str, Any] = None):
        super().__init__(name, config)
        self.url = url
        self.api_config = {
            'method': config.get('method', 'GET'),
            'headers': config.get('headers', {}),
            'params': config.get('params', {}),
            'timeout': config.get('timeout', 30),
            'auth': config.get('auth', None),
            'json_path': config.get('json_path', None)
        }
    
    def extract(self) -> pd.DataFrame:
        """Extract data from API"""
        try:
            # Make API request
            response = requests.request(
                method=self.api_config['method'],
                url=self.url,
                headers=self.api_config['headers'],
                params=self.api_config['params'],
                timeout=self.api_config['timeout'],
                auth=self.api_config['auth']
            )
            
            response.raise_for_status()
            data = response.json()
            
            # Extract specific path from JSON if specified
            if self.api_config['json_path']:
                for key in self.api_config['json_path'].split('.'):
                    data = data[key]
            
            # Convert to DataFrame
            df = pd.json_normalize(data)
            
            self.logger.info(f"Extracted {len(df)} rows from API: {self.url}")
            return df
            
        except Exception as e:
            raise ETLPipelineError(f"Failed to extract from API {self.url}: {e}")

class DirectoryScanner(DataExtractor):
    """Scan directory for files and extract metadata"""
    
    def __init__(self, name: str, directory_path: str, config: Dict[str, Any] = None):
        super().__init__(name, config)
        self.directory_path = directory_path
        self.scan_config = {
            'file_pattern': config.get('file_pattern', '*'),
            'recursive': config.get('recursive', False),
            'include_hidden': config.get('include_hidden', False),
            'extract_metadata': config.get('extract_metadata', True)
        }
    
    def extract(self) -> pd.DataFrame:
        """Scan directory and extract file information"""
        if not os.path.exists(self.directory_path):
            raise ETLPipelineError(f"Directory not found: {self.directory_path}")
        
        try:
            file_data = []
            directory = Path(self.directory_path)
            
            # Get files based on pattern
            if self.scan_config['recursive']:
                files = directory.rglob(self.scan_config['file_pattern'])
            else:
                files = directory.glob(self.scan_config['file_pattern'])
            
            for file_path in files:
                if file_path.is_file():
                    # Skip hidden files if not included
                    if not self.scan_config['include_hidden'] and file_path.name.startswith('.'):
                        continue
                    
                    file_info = {
                        'file_path': str(file_path),
                        'file_name': file_path.name,
                        'file_extension': file_path.suffix,
                        'directory': str(file_path.parent),
                        'file_size': file_path.stat().st_size,
                        'created_time': datetime.fromtimestamp(file_path.stat().st_ctime),
                        'modified_time': datetime.fromtimestamp(file_path.stat().st_mtime)
                    }
                    
                    # Extract additional metadata if requested
                    if self.scan_config['extract_metadata']:
                        file_info.update(self._extract_file_metadata(file_path))
                    
                    file_data.append(file_info)
            
            df = pd.DataFrame(file_data)
            self.logger.info(f"Scanned {len(df)} files from directory: {self.directory_path}")
            return df
            
        except Exception as e:
            raise ETLPipelineError(f"Failed to scan directory {self.directory_path}: {e}")
    
    def _extract_file_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Extract additional metadata from file"""
        metadata = {}
        
        try:
            # Basic file stats
            stat = file_path.stat()
            metadata.update({
                'access_time': datetime.fromtimestamp(stat.st_atime),
                'permissions': oct(stat.st_mode)[-3:],
                'owner_id': stat.st_uid,
                'group_id': stat.st_gid
            })
            
            # File type specific metadata
            if file_path.suffix.lower() in ['.csv', '.txt']:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        line_count = sum(1 for _ in f)
                    metadata['line_count'] = line_count
                except:
                    metadata['line_count'] = None
            
            elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                try:
                    workbook = openpyxl.load_workbook(file_path, read_only=True)
                    metadata['sheet_count'] = len(workbook.sheetnames)
                    metadata['sheet_names'] = workbook.sheetnames
                    workbook.close()
                except:
                    metadata['sheet_count'] = None
                    metadata['sheet_names'] = None
            
        except Exception as e:
            self.logger.warning(f"Failed to extract metadata for {file_path}: {e}")
        
        return metadata

class LibraryDataExtractor(DataExtractor):
    """Specialized extractor for library system data"""
    
    def __init__(self, name: str, db_path: str, config: Dict[str, Any] = None):
        super().__init__(name, config)
        self.db_path = db_path
        self.extraction_type = config.get('extraction_type', 'full')
        self.date_range = config.get('date_range', {})
    
    def extract(self) -> pd.DataFrame:
        """Extract library-specific data"""
        try:
            conn = sqlite3.connect(self.db_path)
            
            if self.extraction_type == 'books':
                query = """
                SELECT b.*, a.name as author_name, p.name as publisher_name
                FROM books b
                LEFT JOIN authors a ON b.author_id = a.id
                LEFT JOIN publishers p ON b.publisher_id = p.id
                """
                
            elif self.extraction_type == 'transactions':
                query = """
                SELECT t.*, m.name as member_name, b.title as book_title
                FROM transactions t
                LEFT JOIN members m ON t.member_id = m.id
                LEFT JOIN books b ON t.book_id = b.id
                """
                
                # Add date filter if specified
                if self.date_range:
                    start_date = self.date_range.get('start_date')
                    end_date = self.date_range.get('end_date')
                    
                    if start_date:
                        query += f" WHERE t.issue_date >= '{start_date}'"
                    if end_date:
                        if start_date:
                            query += f" AND t.issue_date <= '{end_date}'"
                        else:
                            query += f" WHERE t.issue_date <= '{end_date}'"
                
            elif self.extraction_type == 'members':
                query = """
                SELECT m.*, COUNT(t.id) as total_transactions
                FROM members m
                LEFT JOIN transactions t ON m.id = t.member_id
                GROUP BY m.id
                """
                
            elif self.extraction_type == 'analytics':
                query = """
                SELECT 
                    DATE(t.issue_date) as date,
                    COUNT(*) as daily_transactions,
                    COUNT(DISTINCT t.member_id) as active_members,
                    COUNT(DISTINCT t.book_id) as books_issued
                FROM transactions t
                GROUP BY DATE(t.issue_date)
                ORDER BY date
                """
                
            else:
                # Full extraction - join all main tables
                query = """
                SELECT 
                    t.id as transaction_id,
                    t.issue_date,
                    t.return_date,
                    t.status,
                    m.name as member_name,
                    m.email as member_email,
                    m.phone as member_phone,
                    b.title as book_title,
                    b.isbn,
                    b.genre,
                    a.name as author_name,
                    p.name as publisher_name
                FROM transactions t
                JOIN members m ON t.member_id = m.id
                JOIN books b ON t.book_id = b.id
                LEFT JOIN authors a ON b.author_id = a.id
                LEFT JOIN publishers p ON b.publisher_id = p.id
                """
            
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            self.logger.info(f"Extracted {len(df)} library records ({self.extraction_type})")
            return df
            
        except Exception as e:
            raise ETLPipelineError(f"Failed to extract library data: {e}")

# Example usage and configuration
EXTRACTOR_CONFIGS = {
    'library_books': {
        'class': LibraryDataExtractor,
        'config': {
            'extraction_type': 'books'
        }
    },
    'library_transactions': {
        'class': LibraryDataExtractor,
        'config': {
            'extraction_type': 'transactions',
            'date_range': {
                'start_date': '2024-01-01',
                'end_date': '2024-12-31'
            }
        }
    },
    'csv_data': {
        'class': CSVExtractor,
        'config': {
            'encoding': 'utf-8',
            'delimiter': ',',
            'header': 0
        }
    },
    'excel_data': {
        'class': ExcelExtractor,
        'config': {
            'sheet_name': 0,
            'header': 0,
            'engine': 'openpyxl'
        }
    }
}

if __name__ == "__main__":
    print("📊 Data Extraction Components Ready")
    print("Available extractors:")
    for name, config in EXTRACTOR_CONFIGS.items():
        print(f"  - {name}: {config['class'].__name__}")
